"""Python library extending yEd functionality through programmatic interface to graphs.

Currently primarily supports building of yEd graph objects from scratch.

Planned is the addition of read, management and other capabilities.

"""

import os
import platform
import re
import subprocess
import sys
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field  # FIXME: to be incorporated
from shutil import which
from tkinter import messagebox as msg
from typing import Any, List, Optional
from xml.dom import minidom

import psutil
import pygetwindow as gw
from openpyxl import Workbook


# Enumerated parameters / Constants
PROGRAM_NAME = "yEd.exe"

LINE_TYPES = [
    "line",
    "dashed",
    "dotted",
    "dashed_dotted",
]

FONT_STYLES = [
    "plain",
    "bold",
    "italic",
    "bolditalic",
]

HORIZONTAL_ALIGNMENTS = [
    "left",
    "center",
    "right",
]

VERTICAL_ALIGNMENTS = [
    "top",
    "center",
    "bottom",
]

CUSTOM_PROPERTY_SCOPES = [
    "node",
    "edge",
]  # TODO: DOES THIS NEED GROUP?

CUSTOM_PROPERTY_TYPES = [
    "string",
    "int",
    "double",
    "boolean",
]


class InvalidValueError(ValueError):
    pass  # FIXME:


def checkValue(
    parameter_name: str,
    value: Any,
    validValues: Optional[List[str]] = None,
) -> None:
    """Check whether given graph property inputs
    (e.g. Shape, Arrow type, Line type, etc.)
    are valid (within existing enumerated options).
    If not valid - returns error message for input."""

    if validValues:
        if value not in validValues:
            raise InvalidValueError(f"{parameter_name} '{value}' is not supported. Use: '{', '.join(validValues)}'")


class Graph_file:
    def __init__(self, file_name_or_path=None):
        self.DEFAULT_FILE_NAME = "temp"
        self.EXTENSION = ".graphml"
        self.dir = self.path_validate(file_name_or_path)
        self.basename = self.base_name_validate(file_name_or_path)
        self.fullpath = os.path.join(self.dir, self.basename)
        self.window_search_name = self.basename + " - yEd"
        self.file_exists = os.path.isfile(self.fullpath)

    def path_validate(self, temp_name_or_path=None):
        path = ""
        if temp_name_or_path:
            path = os.path.dirname(temp_name_or_path)
        return path or os.getcwd()

    def base_name_validate(self, temp_name_or_path=None):
        temp_name = ""
        if temp_name_or_path:
            temp_name = os.path.basename(temp_name_or_path)
        temp_name = temp_name or f"{self.DEFAULT_FILE_NAME}"
        if not temp_name.endswith(self.EXTENSION):
            temp_name += self.EXTENSION
        return temp_name

    def open_with_yed(self, force=False):
        print("opening...")
        open_yed_file(self, force)


class Label:
    """Generic Label Class for nodes / edges in yEd"""

    graphML_tagName = None

    def __init__(
        self,
        text="",
        height="18.1328125",
        width=None,
        alignment="center",
        font_family="Dialog",
        font_size="12",
        font_style="plain",
        horizontalTextPosition="center",
        underlined_text="false",
        text_color="#000000",
        icon_text_gap="4",
        horizontal_text_position="center",
        vertical_text_position="center",
        visible="true",
        border_color=None,
        background_color=None,
        has_background_color="false",
    ):
        # make class abstract
        if type(self) is Label:
            raise Exception("Label is an abstract class and cannot be instantiated directly")

        self._text = text

        # Initialize dictionary for parameters
        self._params = {}
        self.updateParam("horizontalTextPosition", horizontal_text_position, HORIZONTAL_ALIGNMENTS)
        self.updateParam("verticalTextPosition", vertical_text_position, VERTICAL_ALIGNMENTS)
        self.updateParam("alignment", alignment, HORIZONTAL_ALIGNMENTS)
        self.updateParam("fontStyle", font_style, FONT_STYLES)

        # TODO: Implement range checks
        self.updateParam("fontFamily", font_family)
        self.updateParam("iconTextGap", icon_text_gap)
        self.updateParam("fontSize", font_size)
        self.updateParam("textColor", text_color)
        self.updateParam("visible", visible.lower(), ["true", "false"])
        self.updateParam("underlinedText", underlined_text.lower(), ["true", "false"])
        if background_color:
            has_background_color = "true"
        self.updateParam("hasBackgroundColor", has_background_color.lower(), ["true", "false"])
        self.updateParam("width", width)
        self.updateParam("height", height)
        self.updateParam("borderColor", border_color)
        self.updateParam("backgroundColor", background_color)

    def updateParam(
        self,
        parameter_name,
        value,
        validValues=None,
    ):
        if value is None:
            return False
        checkValue(parameter_name, value, validValues)

        self._params[parameter_name] = value
        return True

    def addSubElement(self, shape):
        label = ET.SubElement(shape, self.graphML_tagName, **self._params)
        label.text = self._text


class NodeLabel(Label):
    """Node specific label"""

    VALIDMODELPARAMS = {
        "internal": ["t", "b", "c", "l", "r", "tl", "tr", "bl", "br"],
        "corners": ["nw", "ne", "sw", "se"],
        "sandwich": ["n", "s"],
        "sides": ["n", "e", "s", "w"],
        "eight_pos": ["n", "e", "s", "w", "nw", "ne", "sw", "se"],
    }

    graphML_tagName = "y:NodeLabel"

    def __init__(
        self,
        text,
        alignment="center",
        font_family="Dialog",
        font_size="12",
        font_style="plain",
        height="18.1328125",
        horizontalTextPosition="center",
        underlined_text="false",
        icon_text_gap="4",
        text_color="#000000",
        horizontal_text_position="center",
        vertical_text_position="center",
        visible="true",
        has_background_color="false",
        width="55.708984375",
        model_name="internal",
        border_color=None,
        background_color=None,
        model_position="c",
    ):
        super().__init__(
            text,
            height,
            width,
            alignment,
            font_family,
            font_size,
            font_style,
            horizontalTextPosition,
            underlined_text,
            text_color,
            icon_text_gap,
            horizontal_text_position,
            vertical_text_position,
            visible,
            border_color,
            background_color,
            has_background_color,
        )

        self.updateParam("modelName", model_name, NodeLabel.VALIDMODELPARAMS.keys())
        self.updateParam("modelPosition", model_position, NodeLabel.VALIDMODELPARAMS[model_name])


class EdgeLabel(Label):
    """Edge specific label"""

    VALIDMODELPARAMS = {
        "two_pos": ["head", "tail"],
        "centered": ["center"],
        "six_pos": ["shead", "thead", "head", "stail", "ttail", "tail"],
        "three_center": ["center", "scentr", "tcentr"],
        "center_slider": None,
        "side_slider": None,
    }

    graphML_tagName = "y:EdgeLabel"

    def __init__(
        self,
        text,
        alignment="center",
        font_family="Dialog",
        font_size="12",
        font_style="plain",
        height="18.1328125",
        horizontalTextPosition="center",
        underlined_text="false",
        icon_text_gap="4",
        text_color="#000000",
        horizontal_text_position="center",
        vertical_text_position="center",
        visible="true",
        has_background_color="false",
        width="55.708984375",
        model_name="centered",
        model_position="center",
        border_color=None,
        background_color=None,
        preferred_placement=None,
    ):
        super().__init__(
            text,
            height,
            width,
            alignment,
            font_family,
            font_size,
            font_style,
            horizontalTextPosition,
            underlined_text,
            text_color,
            icon_text_gap,
            horizontal_text_position,
            vertical_text_position,
            visible,
            border_color,
            background_color,
            has_background_color,
        )

        self.updateParam("modelName", model_name, EdgeLabel.VALIDMODELPARAMS.keys())
        self.updateParam("modelPosition", model_position, EdgeLabel.VALIDMODELPARAMS[model_name])
        self.updateParam("preferredPlacement", preferred_placement)


class CustomPropertyDefinition:
    """Custom properties which can be added to yEd objects / graph as a whole"""

    def __init__(
        self,
        scope,
        name,
        property_type,
        default_value,
    ):
        """
        scope: [node|edge]
        name: name of the custom property
        property_type: [string|boolean|int|double]
                        boolean: Java keywords [true|false]
        default_value: any above datatype represented as a string
        """
        self.scope = scope
        self.name = name
        self.property_type = property_type
        self.default_value = default_value
        self.id = "%s_%s" % (self.scope, self.name)

    def convert_to_xml(self):
        custom_prop_key = ET.Element("key", id=self.id)
        custom_prop_key.set("for", self.scope)
        custom_prop_key.set("attr.name", self.name)
        custom_prop_key.set("attr.type", self.property_type)

        return custom_prop_key


class Group:
    """yEd Group Object (Visual Container of Nodes / Edges / also can recursively act as Node)"""

    VALID_SHAPES = [
        "rectangle",
        "rectangle3d",
        "roundrectangle",
        "diamond",
        "ellipse",
        "fatarrow",
        "fatarrow2",
        "hexagon",
        "octagon",
        "parallelogram",
        "parallelogram2",
        "star5",
        "star6",
        "star6",
        "star8",
        "trapezoid",
        "trapezoid2",
        "triangle",
        "trapezoid2",
        "triangle",
    ]

    def __init__(
        self,
        group_id,
        parent_graph,
        label=None,
        label_alignment="center",
        shape="rectangle",
        closed="false",
        font_family="Dialog",
        underlined_text="false",
        font_style="plain",
        font_size="12",
        fill="#FFCC00",
        transparent="false",
        border_color="#000000",
        border_type="line",
        border_width="1.0",
        height=False,
        width=False,
        x=False,
        y=False,
        custom_properties=None,
        description="",
        url="",
    ):
        self.label = label
        if label is None:
            self.label = group_id

        self.parent = None
        self.group_id = group_id
        self.nodes = {}
        self.groups = {}
        self.parent_graph = parent_graph
        self.edges = {}
        self.num_edges = 0

        # node shape
        checkValue("shape", shape, Group.VALID_SHAPES)
        self.shape = shape

        self.closed = closed

        # label formatting options
        self.font_family = font_family
        self.underlined_text = underlined_text

        checkValue("font_style", font_style, FONT_STYLES)
        self.font_style = font_style
        self.font_size = font_size

        checkValue("label_alignment", label_alignment, HORIZONTAL_ALIGNMENTS)
        self.label_alignment = label_alignment

        self.fill = fill
        self.transparent = transparent

        self.geom = {}
        if height:
            self.geom["height"] = height
        if width:
            self.geom["width"] = width
        if x:
            self.geom["x"] = x
        if y:
            self.geom["y"] = y

        self.border_color = border_color
        self.border_width = border_width

        checkValue("border_type", border_type, LINE_TYPES)
        self.border_type = border_type

        self.description = description
        self.url = url

        # Handle Node Custom Properties
        for name, definition in Node.custom_properties_defs.items():
            if custom_properties:
                for k, v in custom_properties.items():
                    if k not in Node.custom_properties_defs:
                        raise RuntimeWarning("key %s not recognised" % k)
                    if name == k:
                        setattr(self, name, custom_properties[k])
                        break
                else:
                    setattr(self, name, definition.default_value)
            else:
                setattr(self, name, definition.default_value)

    def add_node(self, node_name, **kwargs):
        """Adding node within Group"""
        if node_name in self.parent_graph.existing_entities:
            raise RuntimeWarning("Node %s already exists" % node_name)

        node = Node(node_name, **kwargs)
        node.parent = self
        self.nodes[node_name] = node
        self.parent_graph.existing_entities[node_name] = node
        return node

    def add_group(self, group_id, **kwargs):
        """Adding a group within current group object (and same parent graph)."""
        if group_id in self.parent_graph.existing_entities:
            raise RuntimeWarning("Node %s already exists" % group_id)

        group = Group(group_id, self.parent_graph, **kwargs)
        group.parent = self
        self.groups[group_id] = group
        self.parent_graph.existing_entities[group_id] = group
        return group

    def is_ancestor(self, node):
        """Check for possible nesting conflict of this id usage"""
        return node.parent is not None and (node.parent is self or self.is_ancestor(node.parent))

    def add_edge(self, node1_name, node2_name, **kwargs):
        """Adds edge - input: node names, not actual node objects"""

        node1 = self.parent_graph.existing_entities.get(node1_name) or self.add_node(node1_name)

        node2 = self.parent_graph.existing_entities.get(node2_name) or self.add_node(node2_name)

        # http://graphml.graphdrawing.org/primer/graphml-primer.html#Nested
        # The edges between two nodes in a nested graph have to be declared in a graph,
        # which is an ancestor of both nodes in the hierarchy.

        if not (self.is_ancestor(node1) and self.is_ancestor(node2)):
            raise RuntimeWarning("Group %s is not ancestor of both %s and %s" % (self.group_id, node1_name, node2_name))

        self.parent_graph.num_edges += 1
        kwargs["edge_id"] = str(self.parent_graph.num_edges)
        edge = Edge(node1_name, node2_name, **kwargs)
        self.edges[edge.edge_id] = edge
        return edge

    def convert_to_xml(self):
        """Converting graph object to graphml xml object"""

        node = ET.Element("node", id=self.group_id)
        node.set("yfiles.foldertype", "group")
        data = ET.SubElement(node, "data", key="data_node")

        # node for group
        pabn = ET.SubElement(data, "y:ProxyAutoBoundsNode")
        r = ET.SubElement(pabn, "y:Realizers", active="0")
        group_node = ET.SubElement(r, "y:GroupNode")

        if self.geom:
            ET.SubElement(group_node, "y:Geometry", **self.geom)

        ET.SubElement(group_node, "y:Fill", color=self.fill, transparent=self.transparent)

        ET.SubElement(
            group_node,
            "y:BorderStyle",
            color=self.border_color,
            type=self.border_type,
            width=self.border_width,
        )

        label = ET.SubElement(
            group_node,
            "y:NodeLabel",
            modelName="internal",
            modelPosition="t",
            fontFamily=self.font_family,
            fontSize=self.font_size,
            underlinedText=self.underlined_text,
            fontStyle=self.font_style,
            alignment=self.label_alignment,
        )
        label.text = self.label

        ET.SubElement(group_node, "y:Shape", type=self.shape)

        ET.SubElement(group_node, "y:State", closed=self.closed)

        graph = ET.SubElement(node, "graph", edgedefault="directed", id=self.group_id)

        if self.url:
            url_node = ET.SubElement(node, "data", key="url_node")
            url_node.text = self.url

        if self.description:
            description_node = ET.SubElement(node, "data", key="description_node")
            description_node.text = self.description

        for node_id in self.nodes:
            n = self.nodes[node_id].convert_to_xml()
            graph.append(n)

        for group_id in self.groups:
            n = self.groups[group_id].convert_to_xml()
            graph.append(n)

        for edge_id in self.edges:
            e = self.edges[edge_id].convert_to_xml()
            graph.append(e)

        # Node Custom Properties
        for name, definition in Node.custom_properties_defs.items():
            node_custom_prop = ET.SubElement(node, "data", key=definition.id)
            node_custom_prop.text = getattr(self, name)

        return node
        # ProxyAutoBoundsNode crap just draws bar at top of group


class Node:
    """yEd Node object"""

    custom_properties_defs = {}

    VALID_SHAPES = [
        "rectangle",
        "rectangle3d",
        "roundrectangle",
        "diamond",
        "ellipse",
        "fatarrow",
        "fatarrow2",
        "hexagon",
        "octagon",
        "parallelogram",
        "parallelogram2",
        "star5",
        "star6",
        "star6",
        "star8",
        "trapezoid",
        "trapezoid2",
        "triangle",
        "trapezoid2",
        "triangle",
    ]

    def __init__(
        self,
        node_name,
        label=None,
        label_alignment="center",
        shape="rectangle",
        font_family="Dialog",
        underlined_text="false",
        font_style="plain",
        font_size="12",
        shape_fill="#FFCC00",
        transparent="false",
        border_color="#000000",
        border_type="line",
        border_width="1.0",
        height=False,
        width=False,
        x=False,
        y=False,
        node_type="ShapeNode",
        UML=False,
        custom_properties=None,
        description="",
        url="",
    ):
        self.list_of_labels = []  # initialize list of labels
        if label:
            self.add_label(
                label,
                alignment=label_alignment,
                font_family=font_family,
                underlined_text=underlined_text,
                font_style=font_style,
                font_size=font_size,
            )
        else:
            self.add_label(
                node_name,
                alignment=label_alignment,
                font_family=font_family,
                underlined_text=underlined_text,
                font_style=font_style,
                font_size=font_size,
            )

        self.node_name = node_name

        self.node_type = node_type
        self.UML = UML

        self.parent = None

        # node shape
        checkValue("shape", shape, Node.VALID_SHAPES)
        self.shape = shape

        # shape fill
        self.shape_fill = shape_fill
        self.transparent = transparent

        # border options
        self.border_color = border_color
        self.border_width = border_width

        checkValue("border_type", border_type, LINE_TYPES)
        self.border_type = border_type

        # geometry
        self.geom = {}
        if height:
            self.geom["height"] = height
        if width:
            self.geom["width"] = width
        if x:
            self.geom["x"] = x
        if y:
            self.geom["y"] = y

        self.description = description
        self.url = url

        # Handle Node Custom Properties
        for name, definition in Node.custom_properties_defs.items():
            if custom_properties:
                for k, v in custom_properties.items():
                    if k not in Node.custom_properties_defs:
                        raise RuntimeWarning("key %s not recognised" % k)
                    if name == k:
                        setattr(self, name, custom_properties[k])
                        break
                else:
                    setattr(self, name, definition.default_value)
            else:
                setattr(self, name, definition.default_value)

    def add_label(self, label_text, **kwargs):
        """Adding node label"""
        self.list_of_labels.append(NodeLabel(label_text, **kwargs))
        return self

    def convert_to_xml(self):
        """Converting node object to xml object"""

        node = ET.Element("node", id=str(self.node_name))
        data = ET.SubElement(node, "data", key="data_node")
        shape = ET.SubElement(data, "y:" + self.node_type)

        if self.geom:
            ET.SubElement(shape, "y:Geometry", **self.geom)
        # <y:Geometry height="30.0" width="30.0" x="475.0" y="727.0"/>

        ET.SubElement(shape, "y:Fill", color=self.shape_fill, transparent=self.transparent)

        ET.SubElement(
            shape,
            "y:BorderStyle",
            color=self.border_color,
            type=self.border_type,
            width=self.border_width,
        )

        for label in self.list_of_labels:
            label.addSubElement(shape)

        ET.SubElement(shape, "y:Shape", type=self.shape)

        if self.UML:
            UML = ET.SubElement(shape, "y:UML")

            attributes = ET.SubElement(UML, "y:AttributeLabel", type=self.shape)
            attributes.text = self.UML["attributes"]

            methods = ET.SubElement(UML, "y:MethodLabel", type=self.shape)
            methods.text = self.UML["methods"]

            stereotype = self.UML["stereotype"] if "stereotype" in self.UML else ""
            UML.set("stereotype", stereotype)

        if self.url:
            url_node = ET.SubElement(node, "data", key="url_node")
            url_node.text = self.url

        if self.description:
            description_node = ET.SubElement(node, "data", key="description_node")
            description_node.text = self.description

        # Node Custom Properties
        for name, definition in Node.custom_properties_defs.items():
            node_custom_prop = ET.SubElement(node, "data", key=definition.id)
            node_custom_prop.text = getattr(self, name)

        return node

    @classmethod
    def set_custom_properties_defs(cls, custom_property):
        cls.custom_properties_defs[custom_property.name] = custom_property


class Edge:
    """yEd Edge - connecting Nodes or Groups"""

    custom_properties_defs = {}

    ARROW_TYPES = [
        "none",
        "standard",
        "white_delta",
        "diamond",
        "white_diamond",
        "short",
        "plain",
        "concave",
        "convex",
        "circle",
        "transparent_circle",
        "dash",
        "skewed_dash",
        "t_shape",
        "crows_foot_one_mandatory",
        "crows_foot_many_mandatory",
        "crows_foot_many_optional",
        "crows_foot_one",
        "crows_foot_many",
        "crows_foot_optional",
    ]

    def __init__(
        self,
        node1,
        node2,
        label=None,
        arrowhead="standard",
        arrowfoot="none",
        color="#000000",
        line_type="line",
        width="1.0",
        edge_id="",
        label_background_color="",
        label_border_color="",
        source_label=None,
        target_label=None,
        custom_properties=None,
        description="",
        url="",
    ):
        self.node1 = node1
        self.node2 = node2

        self.list_of_labels = []  # initialize list of labels

        if label:
            self.add_label(
                label,
                border_color=label_border_color,
                background_color=label_background_color,
            )

        if not edge_id:
            edge_id = "%s_%s" % (node1, node2)

        self.edge_id = str(edge_id)

        if source_label is not None:
            self.add_label(
                source_label,
                model_name="six_pos",
                model_position="shead",
                preferred_placement="source_on_edge",
                border_color=label_border_color,
                background_color=label_background_color,
            )

        if target_label is not None:
            self.add_label(
                source_label,
                model_name="six_pos",
                model_position="shead",
                preferred_placement="source_on_edge",
                border_color=label_border_color,
                background_color=label_background_color,
            )

        checkValue("arrowhead", arrowhead, Edge.ARROW_TYPES)
        self.arrowhead = arrowhead

        checkValue("arrowfoot", arrowfoot, Edge.ARROW_TYPES)
        self.arrowfoot = arrowfoot

        checkValue("line_type", line_type, LINE_TYPES)
        self.line_type = line_type

        self.color = color
        self.width = width

        self.description = description
        self.url = url

        # Handle Edge Custom Properties
        for name, definition in Edge.custom_properties_defs.items():
            if custom_properties:
                for k, v in custom_properties.items():
                    if k not in Edge.custom_properties_defs:
                        raise RuntimeWarning("key %s not recognised" % k)
                    if name == k:
                        setattr(self, name, custom_properties[k])
                        break
                else:
                    setattr(self, name, definition.default_value)
            else:
                setattr(self, name, definition.default_value)

    def add_label(self, label_text, **kwargs):
        """Adding edge label"""
        self.list_of_labels.append(EdgeLabel(label_text, **kwargs))
        # Enable method chaining
        return self

    def convert_to_xml(self):
        """Converting edge object to xml object"""

        edge = ET.Element("edge", id=str(self.edge_id), source=str(self.node1), target=str(self.node2))

        data = ET.SubElement(edge, "data", key="data_edge")
        pl = ET.SubElement(data, "y:PolyLineEdge")

        ET.SubElement(pl, "y:Arrows", source=self.arrowfoot, target=self.arrowhead)
        ET.SubElement(pl, "y:LineStyle", color=self.color, type=self.line_type, width=self.width)

        for label in self.list_of_labels:
            label.addSubElement(pl)

        if self.url:
            url_edge = ET.SubElement(edge, "data", key="url_edge")
            url_edge.text = self.url

        if self.description:
            description_edge = ET.SubElement(edge, "data", key="description_edge")
            description_edge.text = self.description

        # Edge Custom Properties
        for name, definition in Edge.custom_properties_defs.items():
            edge_custom_prop = ET.SubElement(edge, "data", key=definition.id)
            edge_custom_prop.text = getattr(self, name)

        return edge

    @classmethod
    def set_custom_properties_defs(cls, custom_property):
        cls.custom_properties_defs[custom_property.name] = custom_property


class Graph:
    """Graph structure - carries yEd graph information"""

    def __init__(self, directed="directed", graph_id="G"):
        self.nodes = {}
        self.edges = {}
        self.num_edges = 0

        self.directed = directed
        self.graph_id = graph_id
        self.existing_entities = {self.graph_id: self}

        self.groups = {}

        self.custom_properties = []

        self.graphml: ET.Element  # FIXME:

    def add_node(self, node_name, **kwargs):
        """Adding defined node to graph"""
        if node_name in self.existing_entities:
            raise RuntimeWarning("Node %s already exists" % node_name)

        node = Node(node_name, **kwargs)
        self.nodes[node_name] = node
        self.existing_entities[node_name] = node
        return node

    def add_edge(self, node1_name, node2_name, **kwargs):
        """Adding edge to graph - uses node names not node objects."""

        self.existing_entities.get(node1_name) or self.add_node(node1_name)
        self.existing_entities.get(node2_name) or self.add_node(node2_name)

        self.num_edges += 1
        kwargs["edge_id"] = str(self.num_edges)
        edge = Edge(node1_name, node2_name, **kwargs)
        self.edges[edge.edge_id] = edge
        return edge

    def add_group(self, group_id, **kwargs):
        """Adding group to graph"""
        if group_id in self.existing_entities:
            raise RuntimeWarning("Node %s already exists" % group_id)

        group = Group(group_id, self, **kwargs)
        self.groups[group_id] = group
        self.existing_entities[group_id] = group
        return group

    def define_custom_property(self, scope, name, property_type, default_value):
        """Adding custom properties to graph (which makes them available on the contained objects in yEd)"""
        if scope not in CUSTOM_PROPERTY_SCOPES:
            raise RuntimeWarning("Scope %s not recognised" % scope)
        if property_type not in CUSTOM_PROPERTY_TYPES:
            raise RuntimeWarning("Property Type %s not recognised" % property_type)
        if not isinstance(default_value, str):
            raise RuntimeWarning("default_value %s needs to be a string" % default_value)
        custom_property = CustomPropertyDefinition(scope, name, property_type, default_value)
        self.custom_properties.append(custom_property)
        if scope == "node":
            Node.set_custom_properties_defs(custom_property)
        elif scope == "edge":
            Edge.set_custom_properties_defs(custom_property)

    def construct_graphml(self):
        """Creating template graphml xml structure and then placing all graph items into it."""

        # Creating XML structure in Graphml format
        # xml = ET.Element("?xml", version="1.0", encoding="UTF-8", standalone="no")

        graphml = ET.Element("graphml", xmlns="http://graphml.graphdrawing.org/xmlns")
        graphml.set("xmlns:java", "http://www.yworks.com/xml/yfiles-common/1.0/java")
        graphml.set("xmlns:sys", "http://www.yworks.com/xml/yfiles-common/markup/primitives/2.0")
        graphml.set("xmlns:x", "http://www.yworks.com/xml/yfiles-common/markup/2.0")
        graphml.set("xmlns:xsi", "http://www.w3.org/2001/XMLSchema-instance")
        graphml.set("xmlns:y", "http://www.yworks.com/xml/graphml")
        graphml.set("xmlns:yed", "http://www.yworks.com/xml/yed/3")
        graphml.set(
            "xsi:schemaLocation",
            "http://graphml.graphdrawing.org/xmlns http://www.yworks.com/xml/schema/graphml/1.1/ygraphml.xsd",
        )

        # Adding some implementation specific keys for identifying urls, descriptions
        node_key = ET.SubElement(graphml, "key", id="data_node")
        node_key.set("for", "node")
        node_key.set("yfiles.type", "nodegraphics")

        # Definition: url for Node
        node_key = ET.SubElement(graphml, "key", id="url_node")
        node_key.set("for", "node")
        node_key.set("attr.name", "url")
        node_key.set("attr.type", "string")

        # Definition: description for Node
        node_key = ET.SubElement(graphml, "key", id="description_node")
        node_key.set("for", "node")
        node_key.set("attr.name", "description")
        node_key.set("attr.type", "string")

        # Definition: url for Edge
        node_key = ET.SubElement(graphml, "key", id="url_edge")
        node_key.set("for", "edge")
        node_key.set("attr.name", "url")
        node_key.set("attr.type", "string")

        # Definition: description for Edge
        node_key = ET.SubElement(graphml, "key", id="description_edge")
        node_key.set("for", "edge")
        node_key.set("attr.name", "description")
        node_key.set("attr.type", "string")

        # Definition: Custom Properties for Nodes and Edges
        for prop in self.custom_properties:
            graphml.append(prop.convert_to_xml())

        edge_key = ET.SubElement(graphml, "key", id="data_edge")
        edge_key.set("for", "edge")
        edge_key.set("yfiles.type", "edgegraphics")

        # Graph node containing actual objects
        graph = ET.SubElement(graphml, "graph", edgedefault=self.directed, id=self.graph_id)

        # Convert python graph objects into xml structure
        for node in self.nodes.values():
            graph.append(node.convert_to_xml())

        for node in self.groups.values():
            graph.append(node.convert_to_xml())

        for edge in self.edges.values():
            graph.append(edge.convert_to_xml())

        self.graphml = graphml

    def persist_graph(self, file=None, pretty_print=False, overwrite=False) -> Graph_file:
        """Convert graphml object->xml tree->graphml file.
        Temporary naming used if not given.
        """

        graph_file = Graph_file(file)

        if graph_file.file_exists and not overwrite:
            raise FileExistsError

        self.construct_graphml()

        if pretty_print:
            raw_str = ET.tostring(self.graphml)
            pretty_str = minidom.parseString(raw_str).toprettyxml()
            with open(graph_file.fullpath, "w") as f:
                f.write(pretty_str)
        else:
            tree = ET.ElementTree(self.graphml)
            tree.write(graph_file.fullpath)  # Uses internal method to XML Etree

        return graph_file

    def stringify_graph(self):
        """Stringified version of graph in graphml format"""
        self.construct_graphml()
        # Py2/3 sigh.
        if sys.version_info.major < 3:
            return ET.tostring(self.graphml, encoding="UTF-8")
        else:
            return ET.tostring(self.graphml, encoding="UTF-8").decode()

    # FIXME: CURRENTLY UNDER CONSTRUCTION
    def from_existing_graph(self, file: str | Graph_file):
        """Parse xml of existing/stored graph file into python"""

        # Manage file input ==============================
        if isinstance(file, Graph_file):
            graph_file = file
        else:
            graph_file = Graph_file(file)
        if not graph_file.file_exists:
            raise FileNotFoundError

        # Simplify input into string ==============================
        graph_str = xml_to_simple_string(graph_file.fullpath)

        # Begin XML parsing
        root = ET.fromstring(graph_str)

        # Extract off key information==============================
        all_keys = root.findall("key")
        key_dict = dict()
        for a_key in all_keys:
            sub_key_dict = dict()

            key_id = a_key.attrib.get("id")
            # sub_key_dict["label"] = a_key.attrib.get("for")
            sub_key_dict["attr"] = a_key.attrib.get("attr.name", None)
            # sub_key_dict["label"] = a_key.attrib.get("type")
            key_dict[key_id] = sub_key_dict

        # Get major graph node
        graph_root = root.find("graph")

        # get major graph info
        graph_dir = graph_root.get("edgedefault")
        graph_id = graph_root.get("id")

        # instantiate graph object
        new_graph = Graph(directed=graph_dir, graph_id=graph_id)

        # Parse graph

        def is_group_node(node):
            return "foldertype" in node.attrib

        def process_node(parent, input_node):
            # Get sub nodes of this node (group or graph)
            current_level_nodes = input_node.findall("node")
            current_level_edges = input_node.findall("edge")

            for node in current_level_nodes:
                # normal nodes
                if not is_group_node(node):
                    node_init_dict = dict()

                    # <node id="n1">
                    node_init_dict["node_name"] = node.attrib.get("id", None)

                    data_nodes = node.findall("data")
                    info_node = None
                    for data_node in data_nodes:
                        info_node = data_node.find("GenericNode") or data_node.find("ShapeNode")
                        if info_node is not None:
                            node_init_dict["node_type"] = info_node.tag

                            node_label = info_node.find("NodeLabel")
                            if node_label is not None:
                                node_init_dict["label"] = node_label.text
                                # TODO: PORT REST OF NODELABEL

                            # <Fill color="#FFCC00" transparent="false" />
                            fill = info_node.find("Fill")
                            if fill is not None:
                                node_init_dict["shape_fill"] = fill.get("color")
                                node_init_dict["transparent"] = fill.get("transparent")

                            # <BorderStyle color="#000000" type="line" width="1.0" />
                            border_style = info_node.find("BorderStyle")
                            if border_style is not None:
                                node_init_dict["border_color"] = border_style.get("color")
                                node_init_dict["border_type"] = border_style.get("type")
                                node_init_dict["border_width"] = border_style.get("width")

                            # <Shape type="rectangle" />
                            shape_sub = info_node.find("Shape")
                            if shape_sub is not None:
                                node_init_dict["shape"] = shape_sub.get("type")

                            uml = info_node.find("UML")
                            if uml is not None:
                                node_init_dict["shape"] = uml.get("AttributeLabel")
                            # TODO: THERE IS FURTHER DETAIL TO PARSE HERE under uml
                        else:
                            info = data_node.text
                            if info is not None:
                                info = re.sub(r"<!\[CDATA\[", "", info)  # unneeded schema
                                info = re.sub(r"\]\]>", "", info)  # unneeded schema
                                print("info:", info)

                                the_key = data_node.attrib.get("key")

                                info_type = key_dict[the_key]["attr"]
                                if info_type in ["url", "description"]:
                                    node_init_dict[info_type] = info
                    # Removing empty items
                    node_init_dict = {key: value for (key, value) in node_init_dict.items() if value is not None}
                    # create node
                    parent.add_node(**node_init_dict)

                # group nodes
                # <node id="n2" yfiles.foldertype="group">
                elif is_group_node(node):
                    group_init_dict = dict()

                    # <node id="n1">
                    group_init_dict["group_id"] = node.attrib.get("id", None)

                    # Actual Group Data ===================================
                    data_nodes = node.findall("data")
                    for data_node in data_nodes:
                        proxy = data_node.find("ProxyAutoBoundsNode")
                        if proxy is not None:
                            realizer = proxy.find("Realizers")

                            group_nodes = realizer.findall("GroupNode")  # TODO: When are there multiple?

                            for group_node in group_nodes:
                                geom_node = group_node.find("Geometry")
                                if geom_node is not None:
                                    group_init_dict["height"] = geom_node.attrib.get("height", None)
                                    group_init_dict["width"] = geom_node.attrib.get("width", None)
                                    group_init_dict["x"] = geom_node.attrib.get("x", None)
                                    group_init_dict["y"] = geom_node.attrib.get("y", None)

                                fill_node = group_node.find("Fill")
                                if fill_node is not None:
                                    group_init_dict["fill"] = fill_node.attrib.get("color", None)
                                    group_init_dict["transparent"] = fill_node.attrib.get("transparent", None)

                                borderstyle_node = group_node.find("BorderStyle")
                                if borderstyle_node is not None:
                                    group_init_dict["border_color"] = borderstyle_node.attrib.get("color", None)
                                    group_init_dict["border_type"] = borderstyle_node.attrib.get("type", None)
                                    group_init_dict["border_width"] = borderstyle_node.attrib.get("width", None)

                                nodelabel_node = group_node.find("NodeLabel")
                                if nodelabel_node is not None:
                                    group_init_dict["label"] = nodelabel_node.text
                                    group_init_dict["font_family"] = nodelabel_node.attrib.get("fontFamily", None)
                                    group_init_dict["font_size"] = nodelabel_node.attrib.get("fontSize", None)
                                    group_init_dict["underlined_text"] = nodelabel_node.attrib.get(
                                        "underlinedText", None
                                    )
                                    group_init_dict["font_style"] = nodelabel_node.attrib.get("fontStyle", None)
                                    group_init_dict["label_alignment"] = nodelabel_node.attrib.get("alignment", None)

                                group_shape_node = group_node.find("Shape")
                                if group_shape_node is not None:
                                    group_init_dict["shape"] = group_shape_node.attrib.get("type", None)

                                group_state_node = group_node.find("State")
                                if group_state_node is not None:
                                    group_init_dict["closed"] = group_state_node.attrib.get("closed", None)
                                    # group_init_dict["aaa"] = group_state_node.attrib.get("closedHeight",None)
                                    # group_init_dict["aaaa"] = group_state_node.attrib.get("closedWidth",None)
                                    # group_init_dict["aaaa"] = group_state_node.attrib.get("innerGraphDisplayEnabled",None)

                                # TODO: GATHER THE MULTIPLE GROUP NODES

                                break

                        else:
                            info = data_node.text
                            if info is not None:
                                info = re.sub(r"<!\[CDATA\[", "", info)  # unneeded schema
                                info = re.sub(r"\]\]>", "", info)  # unneeded schema
                                print("info:", info)

                                the_key = data_node.attrib.get("key")

                                info_type = key_dict[the_key]["attr"]
                                if info_type in ["url", "description"]:
                                    group_init_dict[info_type] = info

                    # Group - Graph node # TODO:
                    sub_graph_node = node.find("graph")
                    # graph_id = sub_graph_node.get("id")

                    # Removing empty items
                    group_init_dict = {key: value for (key, value) in group_init_dict.items() if value is not None}

                    # Creating new group
                    new_group = parent.add_group(**group_init_dict)

                    # Recursive processing
                    if sub_graph_node is not None:
                        process_node(parent=new_group, input_node=sub_graph_node)

                    # parse_graph_node(graph_node)

                # unknown node type
                else:
                    raise NotImplementedError

            # edges then establish connections
            for edge_node in current_level_edges:
                edge_init_dict = dict()

                # <node id="n1">
                edge_init_dict["edge_id"] = edge_node.attrib.get(
                    "id", None
                )  # FIXME: THIS LIKELY NEEDS HELP - GROUP IDS FROM YED PRODUCED FILES LIKE n2::n1
                edge_init_dict["node1_name"] = edge_node.attrib.get("source", None)
                edge_init_dict["node2_name"] = edge_node.attrib.get("target", None)

                # <data key="d5">
                data_nodes = edge_node.findall("data")
                for data_node in data_nodes:
                    polylineedge = data_node.find("PolyLineEdge")

                    if polylineedge is not None:
                        # TODO: ADD POSITION MANAGEMENT
                        # path_node = polylineedge.find("Path")
                        # if path_node:
                        #   edge_init_dict["label"] = path_node.attrib.get("sx")
                        #   edge_init_dict["label"] = path_node.attrib.get("sy")
                        #   edge_init_dict["label"] = path_node.attrib.get("tx")
                        #   edge_init_dict["label"] = path_node.attrib.get("ty")

                        linestyle_node = polylineedge.find("LineStyle")
                        if linestyle_node is not None:
                            edge_init_dict["color"] = linestyle_node.attrib.get("color", None)
                            edge_init_dict["line_type"] = linestyle_node.attrib.get("type", None)
                            edge_init_dict["width"] = linestyle_node.attrib.get("width", None)

                        arrows_node = polylineedge.find("Arrows")
                        if arrows_node is not None:
                            edge_init_dict["arrowfoot"] = arrows_node.attrib.get("source", None)
                            edge_init_dict["arrowhead"] = arrows_node.attrib.get("target", None)

                        edgelabel_node = polylineedge.find("EdgeLabel")
                        if edgelabel_node is not None:
                            edge_init_dict["label"] = edgelabel_node.text
                            edge_init_dict["arrowfoot"] = edgelabel_node.attrib.get("source", None)
                            edge_init_dict["arrowhead"] = edgelabel_node.attrib.get("target", None)

                    else:
                        info = data_node.text
                        if info is not None:
                            info = re.sub(r"<!\[CDATA\[", "", info)  # unneeded schema
                            info = re.sub(r"\]\]>", "", info)  # unneeded schema
                            print("info:", info)

                            the_key = data_node.attrib.get("key")

                            info_type = key_dict[the_key]["attr"]
                            if info_type in ["url", "description"]:
                                edge_init_dict[info_type] = info

                # bendstyle_node = polylineedge.find("BendStyle")
                # edge_init_dict["smoothed"] = linestyle_node.attrib.get("smoothed") # FIXME

                # TODO:
                #   CUSTOM PROPERTIES

                # Removing empty items
                edge_init_dict = {key: value for (key, value) in edge_init_dict.items() if value is not None}
                parent.add_edge(**edge_init_dict)

        process_node(parent=new_graph, input_node=graph_root)

        return new_graph

    def manage_graph_data_in_excel(self):  # TODO: UNDER CONSTRUCTION
        """Displaying groups, nodes, edges in list format"""

        # create workbook
        excel_wb = Workbook()

        # Get worksheet
        excel_ws = excel_wb.active

        # Inserting /organizing sheets
        objects_ws = excel_wb.create_sheet("Objects_and_Groups", 0)
        relations_ws = excel_wb.create_sheet("Relations", 1)
        excel_wb.remove(excel_ws)

        # potentially saving
        excel_wb.save("test.xlsx")

        # nodes and groups
        # graph.existing_entities
        # graph.edges

        # gather nodes
        # gather groups
        # gather edges - node1 node2 label id

        # replace vs
        #     pass


# App related functions -------------------------
def is_yed_findable():
    """Find yed exe path locally"""
    path = which(PROGRAM_NAME) or None
    yed_found_bool = path is not None
    if not yed_found_bool:
        msg.showerror(
            title="Could not find yEd application.",
            message="Please install / add yEd to path.",
        )
    return yed_found_bool


def get_yed_pid():
    pid = None
    for process in psutil.process_iter(["name"]):
        if process.info["name"] == PROGRAM_NAME:
            pid = process.pid or None
            break
    return pid


def get_yed_process():
    process = None
    for process_iter in psutil.process_iter(["name"]):
        if process_iter.info["name"] == PROGRAM_NAME:
            process = process_iter
            break
    return process


def is_yed_open() -> bool:
    """Check whether yEd is currently open - windows, linux, os, etc."""
    return get_yed_pid() is not None


def _maximize(file=None) -> None:
    """Maximize/show the yed app (for particular file) - if found."""

    window = get_yed_graph_window_id(file)

    if window:
        window.activate()
        window.show()
        window.maximize()


def get_yed_graph_window_id(file: Graph_file | None):
    """Retrieve handle of yed window containing file"""
    APP_NAME = "yEd"
    window = None
    if not file:
        potential_windows = [title.lower for title in gw.getAllTitles()]
        search_name = [title for title in potential_windows if title.endswith(APP_NAME)][0]
    else:
        search_name = file.window_search_name

    if search_name:
        window = gw.getWindowsWithTitle(search_name)  # FIXME: gives false positive w files of same name / diff path.
    return window[0] if window else None


def is_yed_graph_open(file: Graph_file) -> bool:
    """Check whether yEd is currently open with particular file - windows, linux, os, etc."""
    window = get_yed_graph_window_id(file)
    return window is not None


def open_yed_file(file: Graph_file, force=False) -> None:
    """Opens yed file - also will start yed if not open."""
    if force:
        answer = msg.askokcancel(title="Force yEd App Close", message="Are you ok to force yEd closure?")
        if not answer:
            print("Exiting program")
            exit()
        process = get_yed_process()
        if process:
            process.kill()
    if not is_yed_graph_open(file):
        if platform.system() == "Darwin":  # macOS
            subprocess.call(("open", file.fullpath))
        elif platform.system() == "Windows":  # Windows
            os.startfile(file.fullpath)
        else:  # linux variants
            subprocess.call(("xdg-open", file.fullpath))

        _maximize(file)  # FIXME: Likely not effective


def start_yed() -> None:
    """Starts yEd program (if installed and on path and not already open)."""
    if is_yed_findable():
        if not is_yed_open():
            subprocess.run(PROGRAM_NAME)


# Utilities
def xml_to_simple_string(file_path) -> str:
    try:
        with open(file_path, "r") as graph_file:
            graph_str = graph_file.read()

    except Exception as E:
        print(f"Error: {E}")

    else:
        # Preprocessing of file for ease of parsing
        graph_str = graph_str.replace("\n", " ")  # line returns
        graph_str = graph_str.replace("\r", " ")  # line returns
        graph_str = graph_str.replace("\t", " ")  # tabs
        graph_str = re.sub("<graphml .*?>", "<graphml>", graph_str)  # unneeded schema
        graph_str = graph_str.replace("> <", "><")  # empty text
        graph_str = graph_str.replace("y:", "")  # unneeded namespace prefix
        graph_str = graph_str.replace("xml:", "")  # unneeded namespace prefix
        graph_str = graph_str.replace("yfiles.", "")  # unneeded namespace prefix
        graph_str = re.sub(" {1,}", " ", graph_str)  # reducing redundant spaces
        # print(graph_str) # debug

    return graph_str
